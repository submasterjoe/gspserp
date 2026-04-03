from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


TEMPLATE_VERSION = "1.0"


@dataclass
class QuotationExcelData:
    project_code: str | None
    tax_percent: Decimal
    valid_until: date | None
    notes: str | None
    lines: list[tuple[str, Decimal, Decimal]]


def build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(color="FFFFFF", bold=True)

    ws["A1"] = "GSPS Quotation Upload Template"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Template version: {TEMPLATE_VERSION}"
    ws["A4"] = "Fill the cells below, then upload the file in the web app."
    ws["A5"] = "Required: at least 1 line item with Description."
    ws["A6"] = "Notes: project_code is optional if you select project in the UI."

    # Meta block
    ws["A8"] = "project_code"
    ws["B8"] = ""
    ws["A9"] = "tax_percent"
    ws["B9"] = 0
    ws["A10"] = "valid_until (YYYY-MM-DD)"
    ws["B10"] = ""
    ws["A11"] = "notes"
    ws["B11"] = ""

    ws["A13"] = "LINES (start from row 14)"
    ws["A13"].font = Font(bold=True)

    cols = ["Description", "Quantity", "UnitPrice"]
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=14, column=i, value=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Example row
    ws["A15"] = "Example: Site supervision"
    ws["B15"] = 120
    ws["C15"] = 800

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A15"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_upload_xlsx(content: bytes) -> QuotationExcelData:
    wb = load_workbook(BytesIO(content), data_only=True)
    if "Quotation" not in wb.sheetnames:
        raise ValueError("Missing sheet 'Quotation'")
    ws = wb["Quotation"]

    def s(cell: str) -> str:
        v = ws[cell].value
        return str(v).strip() if v is not None else ""

    project_code = s("B8") or None
    tax_raw = s("B9") or "0"
    try:
        tax_percent = Decimal(str(tax_raw))
    except Exception:
        tax_percent = Decimal("0")

    vu = None
    vu_s = s("B10")
    if vu_s:
        try:
            vu = date.fromisoformat(vu_s)
        except Exception:
            vu = None

    notes = s("B11") or None

    lines: list[tuple[str, Decimal, Decimal]] = []
    row = 15
    # read down until blank description for several rows
    blank_run = 0
    while row < 500:
        desc = ws.cell(row=row, column=1).value
        desc_s = str(desc).strip() if desc is not None else ""
        if not desc_s:
            blank_run += 1
            if blank_run >= 5:
                break
            row += 1
            continue
        blank_run = 0
        qty_v = ws.cell(row=row, column=2).value
        price_v = ws.cell(row=row, column=3).value
        try:
            qty = Decimal(str(qty_v if qty_v is not None else "1"))
        except Exception:
            qty = Decimal("1")
        try:
            price = Decimal(str(price_v if price_v is not None else "0"))
        except Exception:
            price = Decimal("0")
        lines.append((desc_s, qty, price))
        row += 1

    if not lines:
        raise ValueError("No line items found in Excel (Description column is empty).")

    return QuotationExcelData(
        project_code=project_code,
        tax_percent=tax_percent,
        valid_until=vu,
        notes=notes,
        lines=lines,
    )

