from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


TEMPLATE_VERSION = "1.0"


@dataclass
class SiteRow:
    project_code: str
    name: str
    address: str | None
    lat: Decimal | None
    lng: Decimal | None
    status: str | None
    notes: str | None


def build_sites_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sites"

    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(color="FFFFFF", bold=True)

    ws["A1"] = "GSPS Sites Upload Template"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Template version: {TEMPLATE_VERSION}"
    ws["A4"] = "Fill rows starting from row 6, then upload in the web app."
    ws["A5"] = "Required: ProjectCode, SiteName. Optional: Address, Lat, Lng, Status, Notes."

    cols = ["ProjectCode", "SiteName", "Address", "Lat", "Lng", "Status", "Notes"]
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=6, column=i, value=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws["A7"] = "PRJ-AC-2603-0001"
    ws["B7"] = "Site A"
    ws["C7"] = "1 Demo Road"
    ws["D7"] = 1.352083
    ws["E7"] = 103.819836
    ws["F7"] = "in_progress"
    ws["G7"] = "Client kick-off meeting"

    widths = [18, 22, 36, 12, 12, 14, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A7"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_sites_upload_xlsx(content: bytes) -> list[SiteRow]:
    wb = load_workbook(BytesIO(content), data_only=True)
    if "Sites" not in wb.sheetnames:
        raise ValueError("Missing sheet 'Sites'")
    ws = wb["Sites"]

    out: list[SiteRow] = []
    row = 7
    blank_run = 0
    while row < 2000:
        pc = ws.cell(row=row, column=1).value
        nm = ws.cell(row=row, column=2).value
        pc_s = str(pc).strip() if pc is not None else ""
        nm_s = str(nm).strip() if nm is not None else ""
        if not pc_s and not nm_s:
            blank_run += 1
            if blank_run >= 10:
                break
            row += 1
            continue
        blank_run = 0

        addr = ws.cell(row=row, column=3).value
        lat_v = ws.cell(row=row, column=4).value
        lng_v = ws.cell(row=row, column=5).value
        st = ws.cell(row=row, column=6).value
        notes = ws.cell(row=row, column=7).value

        lat = None
        lng = None
        try:
            if lat_v is not None and str(lat_v).strip() != "":
                lat = Decimal(str(lat_v))
        except Exception:
            lat = None
        try:
            if lng_v is not None and str(lng_v).strip() != "":
                lng = Decimal(str(lng_v))
        except Exception:
            lng = None

        out.append(
            SiteRow(
                project_code=pc_s,
                name=nm_s,
                address=str(addr).strip() if addr is not None and str(addr).strip() else None,
                lat=lat,
                lng=lng,
                status=str(st).strip() if st is not None and str(st).strip() else None,
                notes=str(notes).strip() if notes is not None and str(notes).strip() else None,
            )
        )
        row += 1

    if not out:
        raise ValueError("No rows found in Excel.")
    # basic validation
    bad = [r for r in out if not r.project_code or not r.name]
    if bad:
        raise ValueError("Some rows are missing ProjectCode or SiteName.")
    return out

